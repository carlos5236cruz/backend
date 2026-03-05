import pandas as pd
from io import BytesIO
from typing import List
from openpyxl import load_workbook


def read_excel_columns(file_bytes: bytes) -> List[str]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    ws = wb.active
    columns = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None:
            columns.append(str(cell.value).strip())
    wb.close()
    return columns


def read_excel_products(file_bytes: bytes) -> List[dict]:
    df = pd.read_excel(BytesIO(file_bytes), dtype=str)
    df = df.fillna("")
    products = []
    for _, row in df.iterrows():
        product = {}
        for col in df.columns:
            product[str(col).strip()] = str(row[col]).strip() if row[col] else ""
        products.append(product)
    return products


def generate_export_excel(columns: List[str], products_data: List[dict]) -> bytes:
    df = pd.DataFrame(products_data, columns=columns)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Produtos")
    return output.getvalue()
